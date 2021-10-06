from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import InvalidSessionIdException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import expected_conditions
from time import sleep
import openpyxl
import os
import sys


# JanforAsinにて変換したAsinコードから商品の売れ行きを分析したデータをexcelに記載するプログラム 2/2


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


print("抽出用のExcelパスを入力してください\n(Asinの記載のあるもの)")
output_excel_path_input = input()

output_excel_path = output_excel_path_input  # 'D:/Workfolder/AMAZONScraping/output.xlsx'

output_excel_book = openpyxl.load_workbook(filename=output_excel_path)

output_excel_sheet = output_excel_book.active

max_lens = output_excel_sheet.max_row
i = 2
for cell_obj in output_excel_sheet.iter_rows(min_row=2, min_col=1, max_row=max_lens, max_col=1):

    for asins in cell_obj:

        asin = asins.value
        print(asin)
        if not "AssertionError" in asin:
            if not "ASINが存在しません" in asin:

                url = "https://sedori-rank-kz.com/chrome/chart/draw/{}".format(asin)
                # options.add_extension("C:/Users/hidet/workalert/extension/sedorin.crx")
                # options.add_argument('--headless')
                options = Options()
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')
                options.add_argument('log-level=3')
                driver = webdriver.Chrome(resource_path('./driver/chromedriver.exe'), options=options)
                driver.maximize_window()
                res = driver.get(url)

                driver.set_page_load_timeout(10)
                WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)
                sleep(3)
                # print(i)

                ranking_30 = driver.find_element_by_class_name("count-up-ranking-30")
                ranking_90 = driver.find_element_by_class_name("count-up-ranking-90")
                ranking_180 = driver.find_element_by_class_name("count-up-ranking-180")
                ranking_365 = driver.find_element_by_class_name("count-up-ranking-365")
                ranking_5000_30 = driver.find_element_by_class_name("count-up-ranking-5000-30")
                ranking_5000_90 = driver.find_element_by_class_name("count-up-ranking-5000-90")
                ranking_5000_180 = driver.find_element_by_class_name("count-up-ranking-5000-180")
                ranking_5000_365 = driver.find_element_by_class_name("count-up-ranking-5000-365")
                countdown_new_30 = driver.find_element_by_class_name("count-down-new-30")
                countdown_new_90 = driver.find_element_by_class_name("count-down-new-90")
                countdown_new_180 = driver.find_element_by_class_name("count-down-new-180")
                countdown_new_365 = driver.find_element_by_class_name("count-down-new-365")
                countdown_used_30 = driver.find_element_by_class_name("count-down-used-30")
                countdown_used_90 = driver.find_element_by_class_name("count-down-used-90")
                countdown_used_180 = driver.find_element_by_class_name("count-down-used-180")
                countdown_used_365 = driver.find_element_by_class_name("count-down-used-365")
                collector_count_30 = driver.find_element_by_class_name("count-down-collect-30")
                collector_count_90 = driver.find_element_by_class_name("count-down-collect-90")
                collector_count_180 = driver.find_element_by_class_name("count-down-collect-180")
                collector_count_365 = driver.find_element_by_class_name("count-down-collect-365")

                sleep(1)
                output_excel_sheet.cell(column=2, row=i).value = ranking_30.text
                output_excel_sheet.cell(column=7, row=i).value = ranking_90.text
                ##output_excel_sheet.cell(column=1, row=i).value = str(ranking_180)
                ## output_excel_sheet.cell(column=1, row=i).value = str(ranking_365)
                sleep(1)
                output_excel_sheet.cell(column=3, row=i).value = ranking_5000_30.text
                output_excel_sheet.cell(column=8, row=i).value = ranking_5000_90.text
                ## output_excel_sheet.cell(column=1, row=i).value = str(ranking_5000_180
                ## output_excel_sheet.cell(column=1, row=i).value = str(ranking_5000_365)
                sleep(1)
                output_excel_sheet.cell(column=4, row=i).value = countdown_new_30.text
                output_excel_sheet.cell(column=9, row=i).value = countdown_new_90.text
                ##output_excel_sheet.cell(column=1, row=i).value = str(countdown_new_180)
                ##output_excel_sheet.cell(column=1, row=i).value = str(countdown_new_365)
                sleep(1)
                output_excel_sheet.cell(column=5, row=i).value = countdown_used_30.text
                output_excel_sheet.cell(column=10, row=i).value = countdown_used_90.text
                ##  output_excel_sheet.cell(column=1, row=i).value = str(countdown_used_180)
                ##output_excel_sheet.cell(column=1, row=i).value = str(countdown_used_365)
                sleep(1)
                output_excel_sheet.cell(column=6, row=i).value = collector_count_30.text
                output_excel_sheet.cell(column=11, row=i).value = collector_count_90.text
                ##output_excel_sheet.cell(column=1, row=i).value = str(collector_count_180
                ## output_excel_sheet.cell(column=1, row=i).value = str(collector_count_365)

                # ranking = a.driver.find_elements(By.TAG_NAME,"td")
                output_excel_book.save(output_excel_path_input)

                driver.close()

                i += 1

            else:
                i += 1
        else:
            i += 1

print("抽出が終了しました")
