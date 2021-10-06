from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import InvalidSessionIdException
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions
from selenium.common import exceptions
from time import sleep
import openpyxl
import os
import sys


#Excelに記載されているJANコードをAmazonで検索掛けてAsinコードへ変換しExcelに記載するプログラム 1/2

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


url = "https://www.amazon.co.jp/"
print("Excel,検索用のパスを入力して下さい")
excel_input_root = input()
print("Excel,抽出用のパスを入力してください")
excel_output_root = input()
print("Asin抽出を開始します。")

input_excel_path = excel_input_root#'D:/Workfolder/Coconara/AMAZONScraping/input.xlsx'

input_excel_book = openpyxl.load_workbook(filename=input_excel_path)

input_excel_sheet = input_excel_book.active

output_excel_path = excel_output_root#'D:/Workfolder/Coconara/AMAZONScraping/output.xlsx'

output_excel_book = openpyxl.load_workbook(filename=output_excel_path)

output_excel_sheet = output_excel_book.active

options = Options()
options = ChromeOptions()

#options.add_extension("C:/Users/hidet/workalert/extension/sedorin.crx")

#options.add_argument('--headless')

driver = webdriver.Chrome(resource_path('./driver/chromedriver.exe'),options=options) # ブラウザのウィンドウサイズを固定

driver.maximize_window()

driver.get(url)

driver.implicitly_wait(10)
i = 1

sleep(1)
k = 1
for cell_obj in input_excel_sheet.iter_rows(min_row=2, min_col=1, max_row=1000, max_col=1): 
    
    for jans in cell_obj:
    
    
        try:
            
            jan = jans.value
            print(str(i))
            print(jan)

            search_textbox = driver.find_element_by_id("twotabsearchtextbox")
        
            search_textbox.clear()

            search_textbox.send_keys(str(jan))
            sleep(1)
            search_textbox.submit()
            
            sleep(1)


            if driver.find_element_by_class_name("a-link-normal s-no-outline"):
                cur_url = driver.find_element_by_class_name("a-link-normal s-no-outline").get_attribute("href")
                pos = cur_url.find('dp/')
                asin = cur_url[pos+3:pos+13]
                output_excel_sheet.cell(column=1, row=i).value = str(asin)   
                k = 1

          #  if "tps://aax-" in cur_url:
             #   k += 1
            #    cur_url = driver.find_element_by_class_name("a-link-normal s-no-outline").get_attribute("href")[k]
            #    pos = cur_url.find('dp/')
            #    asin = cur_url[pos+3:pos+13]
            #    output_excel_sheet.cell(column=1, row=i).value = str(asin)   

                print(str(asin))   

            #driver.close()
                i +=1
                  # driver.switch_to.window(original_window)
        
            
                sleep(1)

        except AssertionError:

            assertion = "AssertionError"
        
            print(assertion)
            i +=1
            output_excel_sheet.cell(column=1, row=i).value = str(assertion)
            sleep(2)
            continue
            


        except TimeoutException:
            print("読み込みに時間がかかっています")
            sleep(10)
            pass

output_excel_book.save(excel_output_root)

print("抽出が終了しました")
    
        
    #if driver.find_element_by_class_name("a-size-base"):
    #iframeをすべて表示する処理
    
    # if driver.find_element_by_id('ASIN'):

        #elem_base = driver.find_element_by_id('ASIN')
    

        
    # if elem_base:
        #asin = elem_base.get_attribute("value")
        
        
    
    
        

    #assert len(driver.window_handles) == 2
    #for window_handle in driver.window_handles:
        #if window_handle != original_window:
            # driver.switch_to.window(window_handle)
            # break
    
    

    #else:
        #pass
        
        #sleep(1)
    
    

    
    # driver.implicitly_wait(10)
    #if driver.find_element_by_class_name("a-size-base"):
    #iframeをすべて表示する処理
    

    # elem_base = driver.find_element_by_id('ASIN')
        
    #  if elem_base:
        #   asin = elem_base.get_attribute("value")
        
    # else:
        #  print("NG")
#
        #sleep(1)  
    #driver.switch_to.window(original_window)

    #driver.close()
    # print(str(asin))
    #sleep(1)
    
